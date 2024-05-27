from aiogram import F, Router
from aiogram.types import Message, CallbackQuery, ContentType
from aiogram import Router

from api.berths.crud import get_berths, get_berth_columns, create_berth
from api.berths.schemas import BerthCreate
from api.docks.crud import get_docks, get_dock_columns, create_dock
from api.docks.schemas import DockCreate
from api.routes.schemas import RouteCreate
from api.routes.crud import get_routes, get_route_columns, create_route
from api.lots.schemas import LotCreate
from api.lots.crud import get_lots, get_lot_columns, create_lot
from api.trips.schemas import TripCreate
from api.trips.crud import get_trips, get_trip_columns, create_trip
from api.relations_dock_berth.crud import get_relations_dock_berth, get_relations_dock_berth_columns, create_relation_dock_berth
from api.relations_dock_berth.schemas import RelationDockBerthCreate
from api.ships.crud import get_ships, get_ships_columns, create_ship
from api.ships.schemas import ShipCreate
from api.shipowners.crud import get_shipowners, get_shipowners_columns, create_shipowner
from api.shipowners.schemas import ShipownerCreate

from core.models.db_work import db_work
from aiogram.types import FSInputFile
from openpyxl import Workbook
from .messages import messages
from . import keyboards as kb
import os
from io import BytesIO

import pandas as pd

from .bot_init import bot


handlers_router = Router()


@handlers_router.message(F.text == 'Туториал 💡')
async def tutorial(message: Message):
    await message.answer(messages['tutorial'], reply_markup=kb.in_tutorial)


@handlers_router.message(F.text == 'Работа с БД 📁')
async def work_with_db(message: Message):
    await message.answer(messages["db_work"], reply_markup=kb.tables)
    await message.answer("🌊", reply_markup=kb.in_tables_menu)


@handlers_router.message(F.text == 'Назад в меню ↩️')
async def tutorial(message: Message):
    await message.answer(messages['menu'], reply_markup=kb.main)


@handlers_router.message(F.text == 'К таблицам 📋')
async def tutorial(message: Message):
    await message.answer(messages["db_work"], reply_markup=kb.tables)
    await message.answer("🌊", reply_markup=kb.in_tables_menu)


# Добавление документа

@handlers_router.message(F.content_type == ContentType.DOCUMENT)
async def handle_excel(message: Message):
    if message.document.mime_type == 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet':
        await bot.download(message.document.file_id,message.document.file_name)

        file_path = message.document.file_name
        file_name = message.document.file_name.split(".")[0]

        df = pd.read_excel(file_path)

        match file_name:
            case "berths":
                for _, row in df.iterrows():
                    berth_data = {
                        "berth_id": row["berth_id"] if not pd.isna(row["berth_id"]) else None,
                        "berth_number": row["berth_number"] if not pd.isna(row["berth_number"]) else None,
                        "berth_letter": row["berth_letter"] if not pd.isna(row["berth_letter"]) else None,
                    }

                    berth_in = BerthCreate(**berth_data)
                    _ = create_berth(db_work.get_session(), berth_in)
            case "docks":
                for _, row in df.iterrows():
                    dock_data = {
                        "dock_id": row["dock_id"] if not pd.isna(row["dock_id"]) else None,
                        "dock_name": row["dock_name"] if not pd.isna(row["dock_name"]) else None,
                        "berth_count": row["berth_count"] if not pd.isna(row["berth_count"]) else None,
                        "dock_address": row["dock_address"] if not pd.isna(row["dock_address"]) else None,
                        "dock_type": row["dock_type"] if not pd.isna(row["dock_type"]) else None,
                        "exploitation": row["exploitation"] if not pd.isna(row["exploitation"]) else None,
                        "department": row["department"] if not pd.isna(row["department"]) else None,
                        "work_start_time": row["work_start_time"] if not pd.isna(row["work_start_time"]) else None,
                        "work_end_time": row["work_end_time"] if not pd.isna(row["work_end_time"]) else None,
                        "dock_description": row["dock_description"] if not pd.isna(row["dock_description"]) else None,
                        "long": row["long"] if not pd.isna(row["long"]) else None,
                        "lat": row["lat"] if not pd.isna(row["lat"]) else None,
                    }

                    dock_in = DockCreate(**dock_data)
                    _ = create_dock(db_work.get_session(), dock_in)
            case "relations_dock_berth":
                for _, row in df.iterrows():
                    relation_dock_berth_data = {
                        "dock_id": row["dock_id"] if not pd.isna(row["dock_id"]) else None,
                        "berth_id": row["berth_id"] if not pd.isna(row["berth_id"]) else None,
                    }

<<<<<<< HEAD
                    relation_dock_berth_in= RelationDockBerthCreate(**relation_dock_berth_data)
                    _ = create_relation_dock_berth(db_work.get_session(), relation_dock_berth_in)   
            case "routes":
                for _, row in df.iterrows():
                    route_data = {
                        "route_name": row["route_name"] if not pd.isna(row["route_name"]) else None,
                        "route_short_name": row["route_short_name"] if not pd.isna(row["route_short_name"]) else None,
                        "route_active": row["route_active"] if not pd.isna(row["route_active"]) else None,
                        "route_type_id": row["route_type_id"] if not pd.isna(row["route_type_id"]) else None,
                        "route_type_name": row["route_type_name"] if not pd.isna(row["route_type_name"]) else None,
                        "route_travel_time": row["route_travel_time"] if not pd.isna(row["route_travel_time"]) else None,
                        "route_color": row["route_color"] if not pd.isna(row["route_color"]) else None,
                    }
                    route_in = RouteCreate(**route_data)
                    _ = create_route(db_work.get_session(), route_in)
            case "lots":
                for _, row in df.iterrows():
                    lot_data = {
                        "route_id": row["route_id"] if not pd.isna(row["route_id"]) else None,
                        "lot_name": row["lot_name"] if not pd.isna(row["lot_name"]) else None,
                        "lot_active": row["lot_active"] if not pd.isna(row["lot_active"]) else None,
                    }
                    lot_in = LotCreate(**lot_data)
                    _ = create_lot(db_work.get_session(), lot_in)
            case "trips":
                for _, row in df.iterrows():
                    trip_data = {
                        "lot_id": row["lot_id"] if not pd.isna(row["lot_id"]) else None,
                        "route_id": row["route_id"] if not pd.isna(row["route_id"]) else None,
                        "trip_name": row["trip_name"] if not pd.isna(row["trip_name"]) else None,
                        "trip_active": row["trip_active"] if not pd.isna(row["trip_active"]) else None,
                    }
                    trip_in = TripCreate(**trip_data)
                    _ = create_trip(db_work.get_session(), trip_in)
=======
                    relation_dock_berth_in = RelationDockBerthCreate(**relation_dock_berth_data)
                    _ = create_relation_dock_berth(db_work.get_session(), relation_dock_berth_in)

            case "ships":
                for _, row in df.iterrows():
                    ships_data = {
                        "ship_id": row["ship_id"] if not pd.isna(row["ship_id"]) else None,
                        "ship_name": row["ship_name"] if not pd.isna(row["ship_name"]) else None,
                        "ship_class": row["ship_class"] if not pd.isna(row["ship_class"]) else None,
                        "ship_num": row["ship_num"] if not pd.isna(row["ship_num"]) else None,
                        "ship_capacity": row["ship_capacity"] if not pd.isna(row["ship_capacity"]) else None,
                        "ship_description": row["ship_description"] if not pd.isna(row["ship_description"]) else None,
                        "ship_model": row["ship_model"] if not pd.isna(row["ship_model"]) else None,
                        "shipowner_id": row["shipowner_id"] if not pd.isna(row["shipowner_id"]) else None,
                    }

                    ship_in = ShipCreate(**ships_data)
                    _ = create_ship(db_work.get_session(), ship_in)

            case "shipowners":
                for _, row in df.iterrows():
                    shipowner_data = {
                        "shipowner_id": row["shipowner_id"] if not pd.isna(row["shipowner_id"]) else None,
                        "shipowner_name": row["shipowner_name"] if not pd.isna(row["shipowner_name"]) else None,
                        "shipowner_inn": str(int(row["shipowner_inn"])) if not pd.isna(row["shipowner_inn"]) else None,
                        "shipowner_ogrn": str(int(row["shipowner_ogrn"])) if not pd.isna(row["shipowner_ogrn"]) else None,
                        "shipowner_contacts": row["shipowner_contacts"] if not pd.isna(row["shipowner_contacts"]) else None,
                        "shipowner_url": row["shipowner_url"] if not pd.isna(row["shipowner_url"]) else None,
                    }

                    shipowner_in= ShipownerCreate(**shipowner_data)
                    _ = create_shipowner(db_work.get_session(), shipowner_in)
            case _:
                await message.answer("Файл начинаться с названия таблицы")


                
>>>>>>> ad635c01965cae845829891d26ebf58d676a8677


        os.remove(file_path)

        await message.answer(f"Файл {message.document.file_name} успешно загружен и обработан!")
    else:
        await message.answer("Пожалуйста, отправьте файл в формате Excel.")


# BERTHS ПРИЧАЛЫ


@handlers_router.callback_query(F.data == 'berths')
async def berths(callback: Message):
    await callback.message.delete()
    await callback.message.answer(f"Таблица berths (Причалы)\n\nВыберите действие 👇🏻", reply_markup=kb.berths_actions)
    await callback.message.answer(f"⚓", reply_markup=kb.in_table_actions)


@handlers_router.callback_query(F.data == 'add_berths')
async def add_berths_handler(callback: Message):
    await callback.message.answer(f"Внесите файл с данными в формате Excel 👇🏻", reply_markup=kb.in_table_actions)


@handlers_router.callback_query(F.data == 'get_berths')
async def get_berths_handler(callback: Message):

    berths_list = get_berths(db_work.get_session())
    wb = Workbook()
    ws = wb.active

    colums = get_berth_columns()[:len(get_berth_columns()) - 1]
    ws.append(colums)

    for berth in berths_list:
        ws.append([berth.berth_id, berth.berth_number, berth.berth_letter])

    file_path = "berths.xlsx"
    wb.save(file_path)

    await callback.message.answer_document(FSInputFile(file_path), caption="Файл с данными о причалах", reply_markup=kb.in_table_actions)

    os.remove(file_path)



# DOCKS ДОКИ


@handlers_router.callback_query(F.data == 'docks')
async def docks(callback: Message):
    await callback.message.delete()
    await callback.message.answer(f"Таблица docks (Доки)\n\nВыберите действие 👇🏻", reply_markup=kb.docks_actions)
    await callback.message.answer(f"⚓", reply_markup=kb.in_table_actions)


@handlers_router.callback_query(F.data == 'add_docks')
async def add_docks_handler(callback: Message):
    await callback.message.answer(f"Внесите файл с данными в формате Excel 👇🏻", reply_markup=kb.in_table_actions)


@handlers_router.callback_query(F.data == 'get_docks')
async def get_docks_handler(callback: Message):

    docks_list = get_docks(db_work.get_session())
    wb = Workbook()
    ws = wb.active

    colums = get_dock_columns()[:len(get_dock_columns()) - 1]
    ws.append(colums)

    for dock in docks_list:
        ws.append([dock.dock_id, dock.dock_name, dock.berth_count, 
                   dock.dock_address, dock.dock_type, dock.exploitation, 
                   dock.department, dock.work_start_time, dock.work_end_time, 
                   dock.dock_description, dock.long, dock.lat])

    file_path = "docks.xlsx"
    wb.save(file_path)

    await callback.message.answer_document(FSInputFile(file_path), caption="Файл с данными о доках", reply_markup=kb.in_table_actions)

    os.remove(file_path)



# relations_dock_berth


@handlers_router.callback_query(F.data == 'relations_dock_berth')
async def relations_dock_berth(callback: Message):
    await callback.message.delete()
    await callback.message.answer(f"Таблица relations_dock_berth (Привязка причалов к докам)\n\nВыберите действие 👇🏻", reply_markup=kb.relations_dock_berth_actions)
    await callback.message.answer(f"⚓", reply_markup=kb.in_table_actions)


@handlers_router.callback_query(F.data == 'add_relations_dock_berth')
async def add_relations_dock_berth_handler(callback: Message):
    await callback.message.answer(f"Внесите файл с данными в формате Excel 👇🏻", reply_markup=kb.in_table_actions)


@handlers_router.callback_query(F.data == 'get_relations_dock_berth')
async def get_relations_dock_berth_handler(callback: Message):

    relations_dock_berth_list = get_relations_dock_berth(db_work.get_session())
    wb = Workbook()
    ws = wb.active

    colums = get_relations_dock_berth_columns()[:len(get_relations_dock_berth_columns()) - 1]
    ws.append(colums)

    for relation_dock_berth in relations_dock_berth_list:
        ws.append([relation_dock_berth.dock_id, relation_dock_berth.berth_id])

    file_path = "relations_dock_berth.xlsx"
    wb.save(file_path)

    await callback.message.answer_document(FSInputFile(file_path), caption="Файл с данными о связях причалов с доками", reply_markup=kb.in_table_actions)

    os.remove(file_path)



# SHIPS Судна


@handlers_router.callback_query(F.data == 'ships')
async def ships(callback: Message):
    await callback.message.delete()
    await callback.message.answer(f"Таблица ships (Судна)\n\nВыберите действие 👇🏻", reply_markup=kb.ships_actions)
    await callback.message.answer(f"⚓", reply_markup=kb.in_table_actions)


@handlers_router.callback_query(F.data == 'add_ships')
async def add_ships_handler(callback: Message):
    await callback.message.answer(f"Внесите файл с данными в формате Excel 👇🏻", reply_markup=kb.in_table_actions)


@handlers_router.callback_query(F.data == 'get_ships')
async def get_ships_handler(callback: Message):

    ships_list = get_ships(db_work.get_session())
    wb = Workbook()
    ws = wb.active

    colums = get_ships_columns()[:len(get_ships_columns()) - 1]
    ws.append(colums)

    for ship in ships_list:
        ws.append([ship.ship_id, ship.ship_name, ship.ship_class, 
                   ship.ship_num, ship.ship_capacity, ship.ship_description, 
                   ship.ship_model, ship.shipowner_id])

    file_path = "ships.xlsx"
    wb.save(file_path)

    await callback.message.answer_document(FSInputFile(file_path), caption="Файл с данными о суднах", reply_markup=kb.in_table_actions)

    os.remove(file_path)



# SHIPOWNERS Владельцы суден


@handlers_router.callback_query(F.data == 'shipowners')
async def shipowners(callback: Message):
    await callback.message.delete()
    await callback.message.answer(f"Таблица shipowners (Владельцы суден)\n\nВыберите действие 👇🏻", reply_markup=kb.shipowners_actions)
    await callback.message.answer(f"⚓", reply_markup=kb.in_table_actions)


@handlers_router.callback_query(F.data == 'add_shipowners')
async def add_shipowners_handler(callback: Message):
    await callback.message.answer(f"Внесите файл с данными в формате Excel 👇🏻", reply_markup=kb.in_table_actions)


@handlers_router.callback_query(F.data == 'get_shipowners')
async def get_shipowners_handler(callback: Message):

    shipowners_list = get_shipowners(db_work.get_session())
    wb = Workbook()
    ws = wb.active

    colums = get_shipowners_columns()[:len(get_shipowners_columns()) - 1]
    ws.append(colums)

    for shipowner in shipowners_list:
        ws.append([shipowner.shipowner_id, shipowner.shipowner_name, 
                   shipowner.shipowner_inn, shipowner.shipowner_ogrn, 
                   shipowner.shipowner_contacts, shipowner.shipowner_url])

    file_path = "shipowners.xlsx"
    wb.save(file_path)

    await callback.message.answer_document(FSInputFile(file_path), caption="Файл с данными о владельцах суден", reply_markup=kb.in_table_actions)

    os.remove(file_path)